"""
Handler for Excel files (.xlsx, .xls) containing HL7 messages.

Scans all sheets and all cells for content that looks like HL7 messages.
"""
import io
from typing import List

from app.file_handlers.base import BaseFileHandler


class ExcelHandler(BaseFileHandler):

    def can_handle(self, filename: str, content_type: str) -> bool:
        ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        return ext in {".xlsx", ".xls"}

    def extract_messages(self, data: bytes, filename: str) -> List[str]:
        ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ".xlsx"

        if ext == ".xls":
            return self._read_xls(data)
        return self._read_xlsx(data)

    def _read_xlsx(self, data: bytes) -> List[str]:
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl is required to read .xlsx files.")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Could not open Excel file: {exc}") from exc

        messages = []
        ehr_lines = []
        is_ehr = False
        valid_types = {"PATIENT", "ENCOUNTER", "ALLERGY", "DIAGNOSIS", "LAB_ORDER", "LAB_RESULT", "VITAL", "IMMUNIZATION", "INSURANCE", "NK1", "CHIEF_COMPLAINT", "SYMPTOM", "PROCEDURE", "MEDICATION", "CLINICAL_NOTE"}

        for sheet in wb.worksheets:
            sheet_title = sheet.title.strip().upper()
            if sheet_title in valid_types:
                # Table-format sheet named after an EHR type (e.g. sheet "PATIENT" with column headers)
                is_ehr = True
                for row in sheet.iter_rows(values_only=True):
                    if not any(row): continue
                    str_row = [str(c).strip() if c is not None else "" for c in row]
                    if str_row[0].strip().upper() == sheet_title:
                        ehr_lines.append("|".join(str_row))
                    else:
                        ehr_lines.append(sheet_title + "|" + "|".join(str_row))
            else:
                # Generic sheet — check for pipe-delimited EHR lines in cells OR HL7 MSH lines
                pipe_ehr_lines = []
                for row in sheet.iter_rows(values_only=True):
                    for cell_val in row:
                        if cell_val is None:
                            continue
                        text = str(cell_val).strip()
                        if not text:
                            continue
                        record_type = text.split("|")[0].strip().upper()
                        if text.upper().startswith("MSH|"):
                            messages.append(text)
                            break
                        elif record_type in valid_types:
                            pipe_ehr_lines.append(text)
                            break  # one EHR line per row
                if pipe_ehr_lines:
                    is_ehr = True
                    ehr_lines.extend(pipe_ehr_lines)

        if is_ehr and ehr_lines:
            messages.append("\n".join(ehr_lines))

        if not messages:
            raise ValueError(
                "No HL7 or EHR messages found in Excel file. "
                "Cells should contain HL7 messages starting with MSH|, or sheets should be named after EHR segments (PATIENT, ENCOUNTER, etc)."
            )
        return messages

    def _read_xls(self, data: bytes) -> List[str]:
        try:
            import xlrd
        except ImportError:
            raise ValueError("xlrd is required to read .xls files.")

        try:
            wb = xlrd.open_workbook(file_contents=data)
        except Exception as exc:
            raise ValueError(f"Could not open .xls file: {exc}") from exc

        messages = []
        ehr_lines = []
        is_ehr = False
        valid_types = {"PATIENT", "ENCOUNTER", "ALLERGY", "DIAGNOSIS", "LAB_ORDER", "LAB_RESULT", "VITAL", "IMMUNIZATION", "INSURANCE", "NK1", "CHIEF_COMPLAINT", "SYMPTOM", "PROCEDURE", "MEDICATION", "CLINICAL_NOTE"}

        for sheet in wb.sheets():
            sheet_title = str(sheet.name).strip().upper()
            if sheet_title in valid_types:
                is_ehr = True
                for row_idx in range(sheet.nrows):
                    row_vals = []
                    for col_idx in range(sheet.ncols):
                        cell_val = sheet.cell_value(row_idx, col_idx)
                        row_vals.append(str(cell_val).strip() if cell_val else "")
                    if not any(row_vals): continue
                    if row_vals[0].strip().upper() == sheet_title:
                        ehr_lines.append("|".join(row_vals))
                    else:
                        ehr_lines.append(sheet_title + "|" + "|".join(row_vals))
            else:
                pipe_ehr_lines = []
                for row_idx in range(sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        cell_val = sheet.cell_value(row_idx, col_idx)
                        if not cell_val:
                            continue
                        text = str(cell_val).strip()
                        if not text:
                            continue
                        record_type = text.split("|")[0].strip().upper()
                        if text.upper().startswith("MSH|"):
                            messages.append(text)
                            break
                        elif record_type in valid_types:
                            pipe_ehr_lines.append(text)
                            break
                if pipe_ehr_lines:
                    is_ehr = True
                    ehr_lines.extend(pipe_ehr_lines)

        if is_ehr and ehr_lines:
            messages.append("\n".join(ehr_lines))

        if not messages:
            raise ValueError("No HL7 messages found in .xls file. Sheets should be named after EHR segments.")
        return messages
